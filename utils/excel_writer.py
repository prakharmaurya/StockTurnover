from datetime import datetime
import json
import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side, PatternFill
from openpyxl.styles.colors import Color

file = 'MainDataFile.xlsx'


def writer():
    try:
        print('Getting File.')
        wb = xl.load_workbook(file)
        print('Got the file')
    except:
        print('File not found creating new one.')
        wb = xl.Workbook()
        print('File created.')
    finally:
        try:
            print('Getting Sheet')
            ws = wb['Data']
            print('Got the Sheet')
        except:
            ws = init_file(wb=wb, heading=heading,
                           heading1=heading1, note=note, text1=text1)
    print('Writting some data')
    # cell = WriteOnlyCell(ws, value="hello world")
    # ws.append([cell, 3.14, 'apple'])
    # ws.insert_rows(2)

    write_data_to_table(sign='equity-CM-CW-turnover', ws=ws)
    write_data_to_table(sign='cdsl-FII-DD', ws=ws)
    write_data_to_table(sign='bse-equity-CW-TO', ws=ws)

    # for row in range(14, ws.max_row + 1):
    #     # print(row)
    #     for column in "A":  # Here you can add or reduce the columns
    #         cell_name = "{}{}".format(column, row)
    #         # print(cell_name)
    #         print("cell position {} has value {}".format(
    #             cell_name, ws[cell_name].value))

    try:
        print('Saving file')
        wb.save(file)
    except Exception as e:
        print("Error - Failed data to write of error is {}".format(str(e)))


def write_data_to_table(sign, ws):
    directory = './temp_data/'
    if(sign == 'equity-CM-CW-turnover'):
        try:
            with open(directory+'equity-CM-CW-turnover.json') as f:
                data = json.load(f)
            fileDate = data["Date"]
            row = date_filter(ws=ws, fileDate=fileDate)[1:]
            print("Writing data for " + sign + " is for the date {} and row {}".format(
                fileDate, row))

            ws["S"+str(row)].style = 'text1'
            ws["S"+str(row)] = data["Pro Trades"]["Buy Value (Rs. in Crores)"]
            ws["T"+str(row)].style = 'text1'
            ws["T"+str(row)] = data["Pro Trades"]["Sell Value (Rs. in Crores)"]
            ws["U"+str(row)].style = 'text1'
            ws["U"+str(row)] = data["Pro Trades"]["Buy Value (Rs. in Crores)"] - \
                data["Pro Trades"]["Sell Value (Rs. in Crores)"]

            print("1")

            ws["AC"+str(row)].style = 'text1'
            ws["AC"+str(row)] = data["Others Trades"]["Buy Value (Rs. in Crores)"]
            ws["AD"+str(row)].style = 'text1'
            ws["AD"+str(row)] = data["Others Trades"]["Sell Value (Rs. in Crores)"]
            ws["AE"+str(row)].style = 'text1'
            ws["AE"+str(row)] = data["Others Trades"]["Buy Value (Rs. in Crores)"] - \
                data["Others Trades"]["Sell Value (Rs. in Crores)"]

            print("2")

            ws["AM"+str(row)].style = 'text1'
            ws["AM"+str(row)] = data["BNK&DFI"]["Buy Value (Rs. in Crores)"]
            ws["AN"+str(row)].style = 'text1'
            ws["AN"+str(row)] = data["BNK&DFI"]["Sell Value (Rs. in Crores)"]
            ws["AO"+str(row)].style = 'text1'
            ws["AO"+str(row)] = data["BNK&DFI"]["Buy Value (Rs. in Crores)"] - \
                data["BNK&DFI"]["Sell Value (Rs. in Crores)"]
        except Exception as e:
            print("Error - Failed data to write of " +
                  sign + " error is {}".format(str(e)))
        print("Done!!!")

    if(sign == 'cdsl-FII-DD'):
        try:
            with open(directory+'cdsl-FII-DD-table1.json') as f:
                data = json.load(f)
            fileDate = data["Date"]
            row = date_filter(ws=ws, fileDate=fileDate)[1:]
            print("Writing data for " + sign + " is for the date {} and row {}".format(
                fileDate, row))

            ws["D"+str(row)].style = 'text1'
            ws["D"+str(row)] = float(data["Equity"]["Stock Exchange"]
                                     ["Gross Purchases (Rs Crore)"])
            ws["E"+str(row)].style = 'text1'
            ws["E"+str(row)] = float(data["Equity"]
                                     ["Stock Exchange"]["Gross Sales (Rs Crore)"])
            ws["F"+str(row)].style = 'text1'
            ws["F"+str(row)] = float(data["Equity"]
                                     ["Stock Exchange"]["Net Investment (Rs Crore)"])
        except Exception as e:
            print("Error - Failed data to write of " +
                  sign + " error is {}".format(str(e)))
        print("Done!!!")

    if(sign == 'bse-equity-CW-TO'):
        try:
            with open(directory+'bse-equity-CW-TO-table1.json') as f:
                data = json.load(f)
            fileDate = data["Date"]
            row = date_filter(ws=ws, fileDate=fileDate)[1:]
            print("Writing data for " + sign + " is for the date {} and row {}".format(
                fileDate, row))

            ws["I"+str(row)].style = 'text1'
            ws["I"+str(row)] = float(data["Buy Value"].replace(",", ""))
            ws["J"+str(row)].style = 'text1'
            ws["J"+str(row)] = float(data["Sale Value"].replace(",", ""))
            ws["K"+str(row)].style = 'text1'
            ws["K"+str(row)] = float(data["Net Value"].replace(",", ""))

            with open(directory+'bse-equity-CW-TO-table2.json') as f:
                data = json.load(f)
            fileDate = data["Date"]
            row = date_filter(ws=ws, fileDate=fileDate)[1:]
            print('This is for the date {} and row {}'.format(
                fileDate, row))

            ws["N"+str(row)].style = 'text1'
            ws["N"+str(row)] = float(data["Proprietary"]
                                     ["Buy"].replace(",", ""))
            ws["O"+str(row)].style = 'text1'
            ws["O"+str(row)] = float(data["Proprietary"]
                                     ["Sales"].replace(",", ""))
            ws["P"+str(row)].style = 'text1'
            ws["P"+str(row)] = float(data["Proprietary"]
                                     ["Net"].replace(",", ""))

            ws["X"+str(row)].style = 'text1'
            ws["X"+str(row)] = float(data["Clients"]
                                     ["Buy"].replace(",", ""))
            ws["Y"+str(row)].style = 'text1'
            ws["Y"+str(row)] = float(data["Clients"]
                                     ["Sales"].replace(",", ""))
            ws["Z"+str(row)].style = 'text1'
            ws["Z"+str(row)] = float(data["Clients"]
                                     ["Net"].replace(",", ""))

            ws["AH"+str(row)].style = 'text1'
            ws["AH"+str(row)] = float(data["NRI"]["Buy"].replace(",", ""))
            ws["AI"+str(row)].style = 'text1'
            ws["AI"+str(row)] = float(data["NRI"]["Sales"].replace(",", ""))
            ws["AJ"+str(row)].style = 'text1'
            ws["AJ"+str(row)] = float(data["NRI"]["Net"].replace(",", ""))

        except Exception as e:
            print("Error - Failed data to write of " +
                  sign + " error is {}".format(str(e)))
        print("Done!!!")


def date_filter(ws, fileDate):
    start_append = False
    date_array = []
    for row in range(1, ws.max_row + 1):
        # print(row)
        for column in "A":
            cell_name = column+str(row)
            # print(cell_name)
            if(start_append):
                try:
                    datetime.strptime(str(ws[cell_name].value), "%d-%b-%Y")
                    date_array.append(ws[cell_name].value)
                except ValueError as e:
                    # print('date not found')
                    break
            if(ws[cell_name].value == 'Date'):
                date_heading_cell = row
                start_append = True

    try:
        date_index = date_array.index(fileDate)
        return "A"+str(date_heading_cell+1+date_index)
    except ValueError as e:
        pass

    date_array.append(fileDate)
    date_array.sort(key=lambda date: datetime.strptime(date, "%d-%b-%Y"))
    date_index = date_array.index(fileDate)
    return cell_inserter(ws=ws, row=date_heading_cell+1+date_index,
                         style='text1', fileDate=fileDate)

# return cell_inserter(ws=ws, row=row, style='text1', fileDate=fileDate)


def cell_inserter(ws, row, style, fileDate):
    print('Inserting row new row at ' + str(row))
    ws.insert_rows(row)
    cell_str = 'A'+str(row)
    # insert date in cell
    ws[cell_str].style = style
    ws[cell_str] = fileDate
    print('Date inserted')
    # return cell str and from fn
    return cell_str


def init_file(wb, heading, heading1, note, text1):
    print('"Data" Sheet not found creating new one at 0 pos')
    ws = wb.create_sheet('Data', 0)
    # Registreing the styles
    ws['A100'].style = heading
    ws['A100'].style = heading1
    ws['A100'].style = note
    ws['A100'].style = text1

    multi_cell_text(ws, rowRange=["1"], colRange=["H", "I", "J", "K", "L", "M"],
                    text="MONEY FLOW IN EQUITY")
    multi_cell_text(ws, rowRange=["12"], colRange=["D", "E", "F"],
                    text="FII/FPI (BSE + NSE + MCX-SX)")
    multi_cell_text(ws, rowRange=["12"], colRange=["I", "J", "K"],
                    text="DII(BSE + NSE + MCX-SX)")
    multi_cell_text(ws, rowRange=["12"], colRange=["N", "O", "P"],
                    text="Proprietary BSE")
    multi_cell_text(ws, rowRange=["12"], colRange=["S", "T", "U"],
                    text="Proprietary NSE")
    multi_cell_text(ws, rowRange=["12"], colRange=["X", "Y", "Z"],
                    text="RETAIL+HNI (BSE)")
    multi_cell_text(ws, rowRange=["12"], colRange=["AC", "AD", "AE"],
                    text="RETAIL+HNI (NSE)")
    multi_cell_text(ws, rowRange=["12"], colRange=["AH", "AI", "AJ"],
                    text="NRI (BSE)")
    multi_cell_text(ws, rowRange=["12"], colRange=["AM", "AN", "AO"],
                    text="NRI (NSE)")

    ws['A14'].style = heading1
    ws['A14'] = "Date"
    ws['D14'].style = heading1
    ws['D14'] = "Gross Purchases(Rs. Crore)"
    ws['E14'].style = heading1
    ws['E14'] = "Gross Sales (Rs. Crore)"
    ws['F14'].style = heading1
    ws['F14'] = "Net Investment (Rs. Crore)"
    ws['I14'].style = heading1
    ws['I14'] = "Gross Purchases(Rs. Crore)"
    ws['J14'].style = heading1
    ws['J14'] = "Gross Sales (Rs. Crore)"
    ws['K14'].style = heading1
    ws['K14'] = "Net Investment (Rs. Crore)"
    ws['N14'].style = heading1
    ws['N14'] = "Proprietary Buy"
    ws['O14'].style = heading1
    ws['O14'] = "Proprietary Sales"
    ws['P14'].style = heading1
    ws['P14'] = "Proprietary Net"
    ws['S14'].style = heading1
    ws['S14'] = "Proprietary Buy"
    ws['T14'].style = heading1
    ws['T14'] = "Proprietary Sales"
    ws['U14'].style = heading1
    ws['U14'] = "Proprietary Net"
    ws['X14'].style = heading1
    ws['X14'] = "Clients Buy"
    ws['Y14'].style = heading1
    ws['Y14'] = "Clients Sales"
    ws['Z14'].style = heading1
    ws['Z14'] = "Clients Net"
    ws['AC14'].style = heading1
    ws['AC14'] = "Clients Buy"
    ws['AD14'].style = heading1
    ws['AD14'] = "Clients Sales"
    ws['AE14'].style = heading1
    ws['AE14'] = "Clients Net"
    ws['AH14'].style = heading1
    ws['AH14'] = "NRI Buy"
    ws['AI14'].style = heading1
    ws['AI14'] = "NRI Sales"
    ws['AJ14'].style = heading1
    ws['AJ14'] = "NRI Net"
    ws['AM14'].style = heading1
    ws['AM14'] = "NRI Buy"
    ws['AN14'].style = heading1
    ws['AN14'] = "NRI Sales"
    ws['AO14'].style = heading1
    ws['AO14'] = "NRI Net"

    print('"Data" Sheet created')
    return ws


def multi_cell_text(ws, rowRange, colRange, text, style='heading'):
    # ws.merge_cells('F1:J1')
    ws.merge_cells(colRange[0]+rowRange[0]+':'+colRange[-1]+rowRange[-1])
    for r in rowRange:
        for c in colRange:
            ws[c+r].style = style
    ws[colRange[0]+rowRange[0]] = text


# Visit this for std colors https://openpyxl.readthedocs.io/en/stable/styles.html
####  Heading Styles #####
heading = NamedStyle(name="heading")
bd = Side(style="thick", color="008000")
heading.border = Border(left=bd, top=bd, right=bd, bottom=bd)
heading.font = Font(name="TimesNewRoman", bold=True,
                    italic=True, size=14, color="000033")
heading.fill = PatternFill(start_color="00FFFFCC", fill_type="solid")
heading.alignment = Alignment(horizontal="center", vertical="center")

heading1 = NamedStyle(name="heading1")
bd = Side(style="thin", color="000000")
heading1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
heading1.font = Font(name="Calibri", bold=True,
                     italic=True, size=12, color="000000")
heading1.fill = PatternFill(start_color="00FFFFFC", fill_type="solid")
heading1.alignment = Alignment(horizontal="center", vertical="center")

####  Text Styles #####
text1 = NamedStyle(name="text1")
bd = Side(style="thin", color="000000")
text1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
text1.font = Font(name="Calibri", size=11, color="000000")
# text1.fill = PatternFill(start_color="00CCFFFF", fill_type="solid")
text1.alignment = Alignment(horizontal="center")

####  note Styles #####
note = NamedStyle(name="note")
note.font = Font(name="Calibri", size=13, color="FF0000", italic=True)
note.fill = PatternFill(start_color="00CCFFFF", fill_type="solid")
note.alignment = Alignment(horizontal="center")
